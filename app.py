import streamlit as st
import pandas as pd
import io
import json
from datetime import datetime, timedelta, timezone
from ortools.sat.python import cp_model
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components

# --- ลอง Import library สำหรับ Google Sheets ---
try:
    import gspread
    from oauth2client.service_account import ServiceCredentials
    SHEETS_AVAILABLE = True
except ImportError:
    SHEETS_AVAILABLE = False

# ==========================================
# ⚙️ ฟังก์ชันพื้นฐาน (Date & Style)
# ==========================================
def get_thai_date(date_obj):
    thai_months = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", 
                   "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
    thai_days = ["วันจันทร์", "วันอังคาร", "วันพุธ", "วันพฤหัสบดี", "วันศุกร์", "วันเสาร์", "วันอาทิตย์"]
    day_name = thai_days[date_obj.weekday()]
    day = date_obj.day
    month = thai_months[date_obj.month]
    year = date_obj.year + 543 
    return f"{day_name}ที่ {day} {month} {year}"

def get_header_color(t_idx, day_of_week):
    if day_of_week == 'Normal':
        if t_idx in [0, 1, 3, 4, 11, 12]: return 'orange' 
        if t_idx in [2]: return 'yellow'                 
        if t_idx in [5, 6, 9, 10]: return 'pink'         
        if t_idx in [7, 8]: return 'purple'              
        if t_idx in [13, 14, 15]: return 'blue'          
    else: 
        if t_idx in [0, 1, 4, 5, 12, 13]: return 'orange' 
        if t_idx in [2, 3]: return 'yellow'              
        if t_idx in [6, 7, 10, 11]: return 'pink'        
        if t_idx in [8, 9]: return 'purple'              
        if t_idx in [14, 15]: return 'blue'              
    return None

header_color_map = {
    'orange': PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid'),
    'yellow': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'pink': PatternFill(start_color='F8CECC', end_color='F8CECC', fill_type='solid'),
    'purple': PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid'),
    'blue': PatternFill(start_color='DAE8FC', end_color='DAE8FC', fill_type='solid')
}
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ==========================================
# 📊 ส่วนเชื่อมต่อ Google Sheets (Database)
# ==========================================
def connect_to_gsheet():
    if not SHEETS_AVAILABLE:
        return None
    try:
        # ดึง Credentials จาก Streamlit Secrets
        creds_dict = st.secrets["gcp_service_account"]
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        # เปิด Sheet ชื่อ 'Pharmacy_Schedule_DB' (ต้องสร้างทิ้งไว้และ Share สิทธิ์ให้ Email ใน JSON)
        return client.open("Pharmacy_Schedule_DB")
    except Exception:
        return None

def save_to_db(df, date_str):
    sheet_file = connect_to_gsheet()
    if not sheet_file: return False
    try:
        # หาหรือสร้าง Worksheet ประจำวันที่
        try:
            worksheet = sheet_file.worksheet(date_str)
            sheet_file.del_worksheet(worksheet) # ลบของเก่าถ้าซ้ำเพื่อเขียนทับ
        except: pass
        
        worksheet = sheet_file.add_worksheet(title=date_str, rows="100", cols="20")
        worksheet.update([df.columns.values.tolist()] + df.values.tolist())
        return True
    except: return False

def load_from_db(date_str):
    sheet_file = connect_to_gsheet()
    if not sheet_file: return None
    try:
        worksheet = sheet_file.worksheet(date_str)
        data = worksheet.get_all_records()
        return pd.DataFrame(data)
    except: return None

# ==========================================
# 🧠 ฟังก์ชันคำนวณตาราง (AI Logic)
# ==========================================
VALID_TIMES = ["08.30", "09.00", "09.30", "10.00", "10.30", "11.00", "11.30", "12.00",
               "12.30", "13.00", "13.30", "14.00", "14.30", "15.00", "15.30", "16.00", "16.30"]

def time_to_slot(t_str): return VALID_TIMES.index(t_str)

def generate_schedule(DAY_OF_WEEK, LEAVES, CUSTOM_TASKS, PART_TIME, FIX_BREAKS, FIXED_MAIN_TASKS, SICK_PEOPLE, IS_MWF, REF_SCHEDULE=None):
    model = cp_model.CpModel()
    
    ft_pharmacists = ['เต้น', 'แอน', 'แม็ค', 'โบ้ท', 'ไม้เอก', 'กิ๊ฟ', 'ฟอร์จูน', 'มิ้ลค์', 'ริน', 
                      'อ๊อฟฟี่', 'ออย', 'บี', 'มายด์', 'ขิม', 'บีม', 'มิ้น', 'ใบเตย', 'จีน่า', 'ปอนด์']
    pt_pharmacists = [pt['name'] for pt in PART_TIME]
    all_pharmacists = ft_pharmacists + pt_pharmacists
    
    time_slots = [f"{VALID_TIMES[i]}-{VALID_TIMES[i+1]}" for i in range(16)]
    
    dispensing_tasks = ['จ่ายยา_4', 'จ่ายยา_5', 'จ่ายยา_6', 'จ่ายยา_7', 'จ่ายยา_8', 'จ่ายยา_9', 'จ่ายยา_10', 'จ่ายยา_11']
    ver_cpoe_tasks = ['Ver_1', 'Ver_2', 'Ver_3', 'Ver_4', 'Ver_5', 'Ver_6', 'Ver_7', 'Ver_8', 'Ver_9', 'Ver_10']
    ver_ps_tasks = ['PS_1', 'PS_2', 'PS_3', 'PS_4', 'PS_5', 'PS_6', 'PS_7', 'PS_8', 'PS_9', 'PS_10']
    tasks = dispensing_tasks + ver_cpoe_tasks + ver_ps_tasks + ['Match_C', 'Match_C2', 'Matching', 'พัก', 'งานเฉพาะ', 'ลา', 'นอกเวลา', 'ว่าง']
             
    x = {}
    for p in all_pharmacists:
        for t in range(16):
            for task in tasks: x[p, t, task] = model.NewBoolVar(f'x_{p}_{t}_{task}')
            model.AddExactlyOne(x[p, t, task] for task in tasks)
            model.Add(x[p, t, 'ว่าง'] == 0)

    if DAY_OF_WEEK == 'Wed_Fri': break_slots, b_groups = [6, 7, 8, 9, 10, 11], [(6,8), (8,10), (10,12)] 
    else: break_slots, b_groups = [5, 6, 7, 8, 9, 10], [(5,7), (7,9), (9,11)]

    # 1. จัดการวันลา
    active_ft = []
    leave_slots = set()
    half_day_leaves = set() 
    for p in ft_pharmacists:
        if p in LEAVES:
            l_type = LEAVES[p]
            l_range = range(0,16) if l_type == 'ทั้งวัน' else (range(0,9) if l_type == 'เช้า' else range(7,16))
            if l_type != 'ทั้งวัน': 
                half_day_leaves.add(p)
                active_ft.append(p)
            for t in l_range: 
                model.Add(x[p, t, 'ลา'] == 1)
                leave_slots.add((p, t))
        else: active_ft.append(p)

    for p in ft_pharmacists:
        for t in range(16):
            if (p, t) not in leave_slots: model.Add(x[p, t, 'ลา'] == 0)

    # 2. งานเฉพาะราย
    custom_dict_index = {}
    custom_task_slots_count = {p: 0 for p in ft_pharmacists} 
    for (p, start, end), task_name in CUSTOM_TASKS.items():
        s_idx, e_idx = time_to_slot(start), time_to_slot(end)
        for t in range(s_idx, e_idx):
            model.Add(x[p, t, 'งานเฉพาะ'] == 1)
            custom_dict_index[(p, t)] = task_name
            if p in ft_pharmacists: custom_task_slots_count[p] += 1
            
    for p in all_pharmacists:
        for t in range(16):
            if (p, t) not in custom_dict_index: model.Add(x[p, t, 'งานเฉพาะ'] == 0)

    # 2.5 ป่วยห้ามจ่ายยา
    for p in SICK_PEOPLE:
        if p in all_pharmacists:
            for t in range(16):
                for task in dispensing_tasks: model.Add(x[p, t, task] == 0)

    # ล็อกภาระงานหลัก
    for (p, start, end), task_name in FIXED_MAIN_TASKS.items():
        s_idx, e_idx = time_to_slot(start), time_to_slot(end)
        for t in range(s_idx, e_idx): model.Add(x[p, t, task_name] == 1)

    # 3. จัดการ Part-time
    for pt in PART_TIME:
        p = pt['name']
        s_idx, e_idx = time_to_slot(pt['start']), time_to_slot(pt['end'])
        
        my_dispense_allowed = ['จ่ายยา_7', 'จ่ายยา_8']
        if len(PART_TIME) >= 2: my_dispense_allowed.extend(['จ่ายยา_6', 'จ่ายยา_9'])
            
        pt_all_allowed = my_dispense_allowed + ['Matching', 'พัก', 'นอกเวลา']
        
        for t in range(16): 
            if t < s_idx or t >= e_idx: model.Add(x[p, t, 'นอกเวลา'] == 1)
            else: model.Add(x[p, t, 'นอกเวลา'] == 0)
        
        for t in range(max(0, s_idx), min(16, e_idx)):
            model.Add(sum(x[p, t, task] for task in pt_all_allowed) == 1)

        is_group_a = (pt['start'] == "09.30" and pt['end'] in ["16.00", "16.30"])
        is_group_b = (pt['start'] in ["10.30", "11.00", "11.30", "12.00", "12.30"] and pt['end'] in ["16.00", "16.30"])
        is_group_c = (pt['start'] in ["13.00", "13.30"] and pt['end'] in ["16.00", "16.30"])
        is_group_d = (pt['start'] == "09.30" and pt['end'] in ["13.00", "13.30"])
        is_group_e = (pt['start'] == "09.00" and pt['end'] == "14.30")

        if pt['has_break'] and not (is_group_c or is_group_d or is_group_e):
            if s_idx <= 8 and e_idx > 8:
                model.Add(x[p, 8, 'พัก'] == 1) 
                if e_idx > 9: model.Add(x[p, 9, 'Matching'] == 1) 
                    
        for t in range(16):
            if not (pt['has_break'] and not (is_group_c or is_group_d or is_group_e) and t == 8): model.Add(x[p, t, 'พัก'] == 0)

        # โควตาจ่ายยา
        if is_group_a: 
            model.Add(sum(x[p, t, task] for t in range(16) for task in my_dispense_allowed) == 8)
            model.Add(sum(x[p, t, 'จ่ายยา_7'] for t in range(16)) == 4) 
            model.Add(sum(x[p, t, 'จ่ายยา_8'] for t in range(16)) == 4) 
        elif is_group_b: model.Add(sum(x[p, t, task] for t in range(16) for task in my_dispense_allowed) == 6)
        elif is_group_c or is_group_d: model.Add(sum(x[p, t, task] for t in range(16) for task in my_dispense_allowed) == 4)
        elif is_group_e: model.Add(sum(x[p, t, task] for t in range(16) for task in my_dispense_allowed) == 6)

        # กฎช่องสุดท้ายของ PT (V119)
        if len(PART_TIME) <= 2:
            for t in range(max(s_idx, e_idx - 2), e_idx):
                if 0 <= t < 16: model.Add(sum(x[p, t, task] for task in my_dispense_allowed) == 1)

    # 4. จัดการเวลาพักของ FT
    b_group_vars_ft = {0: [], 1: [], 2: []}
    for p in all_pharmacists:
        if p in ft_pharmacists:
            model.Add(sum(x[p, t, 'นอกเวลา'] for t in range(16)) == 0) 
            for t in range(16): model.Add(x[p, t, 'Matching'] == 0)
        
        if p in active_ft:
            model.Add(sum(x[p, t, 'พัก'] for t in range(16)) == 2)
            choices = [model.NewBoolVar(f'choice_{p}_b{i}') for i in range(3)]
            if p in FIX_BREAKS and p in ft_pharmacists:
                req_b = FIX_BREAKS[p]
                for i in range(3): model.Add(choices[i] == (1 if i == req_b else 0))
            else: model.AddExactlyOne(choices) 
            for i in range(3):
                b_group_vars_ft[i].append(choices[i])
                for t in range(*b_groups[i]): model.Add(x[p, t, 'พัก'] == 1).OnlyEnforceIf(choices[i])
            for t in range(16):
                if t not in break_slots: model.Add(x[p, t, 'พัก'] == 0)
        elif p in ft_pharmacists: model.Add(sum(x[p, t, 'พัก'] for t in range(16)) == 0)

    total_active_ft = len(active_ft)
    if total_active_ft > 0:
        for i in range(3):
            model.Add(sum(b_group_vars_ft[i]) <= 7) 
            model.Add(sum(b_group_vars_ft[i]) >= max(0, (total_active_ft // 3) - 1))

    reward_vars = []
    
    # --- 8. กฎควบคุมจำนวนคนต่อหน้าที่ ---
    for t in range(16):
        for task in tasks:
            if task not in ['พัก', 'งานเฉพาะ', 'ลา', 'นอกเวลา', 'ว่าง', 'Matching', 'Match_C2']:
                model.Add(sum(x[p, t, task] for p in all_pharmacists) <= 1)
        model.Add(sum(x[p, t, 'Match_C2'] for p in all_pharmacists) <= 1)

        if t < 2: req_core = ['จ่ายยา_6', 'จ่ายยา_7', 'จ่ายยา_8', 'จ่ายยา_9', 'Ver_1', 'Ver_2', 'Ver_3', 'PS_1', 'Match_C']
        elif t == 2: req_core = ['จ่ายยา_5', 'จ่ายยา_6', 'จ่ายยา_7', 'จ่ายยา_8', 'จ่ายยา_9', 'Ver_1', 'Ver_2', 'Ver_3', 'PS_1', 'Match_C']
        else: req_core = ['จ่ายยา_5', 'จ่ายยา_6', 'จ่ายยา_7', 'จ่ายยา_8', 'จ่ายยา_9', 'จ่ายยา_10', 'Ver_1', 'Ver_2', 'Ver_3', 'PS_1', 'Match_C']
        for task in req_core: model.Add(sum(x[p, t, task] for p in all_pharmacists) == 1)

        # V118.3: บังคับปิดช่อง 4, 5, 10, 11 ในชั่วโมงแรก
        if t < 2:
            for d_close in ['จ่ายยา_4', 'จ่ายยา_5', 'จ่ายยา_10', 'จ่ายยา_11']: model.Add(sum(x[p, t, d_close] for p in all_pharmacists) == 0)

        if t not in break_slots: model.Add(sum(x[p, t, 'PS_2'] for p in all_pharmacists) == 1)
        else:
            ps2_sum = sum(x[p, t, 'PS_2'] for p in all_pharmacists)
            reward_vars.append(ps2_sum * 100000)

    for t in range(16):
        for i in range(2, 10): model.Add(sum(x[p, t, f'PS_{i+1}'] for p in all_pharmacists) <= sum(x[p, t, f'PS_{i}'] for p in all_pharmacists))
        for i in range(4, 10): model.Add(sum(x[p, t, f'Ver_{i+1}'] for p in all_pharmacists) <= sum(x[p, t, f'Ver_{i}'] for p in all_pharmacists))
        # Hierarchy
        model.Add(sum(x[p, t, 'จ่ายยา_8'] for p in all_pharmacists) <= sum(x[p, t, 'จ่ายยา_7'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่ายยา_6'] for p in all_pharmacists) <= sum(x[p, t, 'จ่ายยา_8'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่ายยา_9'] for p in all_pharmacists) <= sum(x[p, t, 'จ่ายยา_6'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่ายยา_5'] for p in all_pharmacists) <= sum(x[p, t, 'จ่ายยา_9'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่ายยา_10'] for p in all_pharmacists) <= sum(x[p, t, 'จ่ายยา_5'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่ายยา_4'] for p in all_pharmacists) <= sum(x[p, t, 'จ่ายยา_10'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่ายยา_11'] for p in all_pharmacists) <= sum(x[p, t, 'จ่ายยา_4'] for p in all_pharmacists))

    # --- 9. กฎเหล็ก ---
    for p in all_pharmacists:
        for t in range(15):
            for cat in [dispensing_tasks]:
                for task1 in cat:
                    for task2 in cat:
                        if task1 != task2: model.AddImplication(x[p, t, task1], x[p, t+1, task2].Not())

    all_work_categories = [dispensing_tasks, ver_cpoe_tasks, ver_ps_tasks, ['Match_C', 'Match_C2']]
    for p in all_pharmacists:
        for cat in all_work_categories:
            for t in range(14): model.Add(sum(x[p, t+k, task] for task in cat for k in range(3)) <= 2)

    for p in pt_pharmacists:
        for t in range(14): model.Add(sum(x[p, t+k, 'Matching'] for k in range(3)) <= 2)

    # ความยุติธรรม FT
    is_disp_7_vars = []
    for p in ft_pharmacists:
        tot_disp = sum(x[p, t, task] for t in range(16) for task in dispensing_tasks)
        over_3hr_var = model.NewBoolVar(f'over_3hr_{p}')
        model.Add(tot_disp <= 6 + over_3hr_var)
        model.Add(tot_disp <= 7) 
        reward_vars.append(over_3hr_var * -500000) 
        is_disp_7_vars.append(over_3hr_var)
        
        # กฎช่อง 7-8
        done_7 = model.NewBoolVar(f'd7_{p}'); model.Add(sum(x[p,t,'จ่ายยา_7'] for t in range(16)) > 0).OnlyEnforceIf(done_7)
        done_8 = model.NewBoolVar(f'd8_{p}'); model.Add(sum(x[p,t,'จ่ายยา_8'] for t in range(16)) > 0).OnlyEnforceIf(done_8)
        model.Add(done_7 + done_8 <= 1)
    model.Add(sum(is_disp_7_vars) <= 2) 

    # 🌟 V120: Stability Reward (โบนัสสำหรับการจัดเหมือนตารางเดิม) 🌟
    if REF_SCHEDULE is not None:
        task_map = {
            "Match + C": "Match_C", "Match + C2": "Match_C2", "Matching": "Matching", "พัก": "พัก", "ลา": "ลา", "-": "นอกเวลา",
            "Ver 1 INC": "Ver_1", "Ver 2/ปณ.": "Ver_2", "Ver 3/A": "Ver_3"
        }
        for p in all_pharmacists:
            if p in REF_SCHEDULE.index:
                for t_idx, t_col in enumerate(time_slots):
                    old_task_display = REF_SCHEDULE.at[p, t_col]
                    # แปลงชื่อกลับเป็น internal
                    target_task = None
                    if "จ่าย " in old_task_display: target_task = old_task_display.replace("จ่าย ", "จ่ายยา_")
                    elif "Ver PS" in old_task_display: target_task = old_task_display.replace("Ver PS", "PS_")
                    elif "Ver " in old_task_display: target_task = old_task_display.replace("Ver ", "Ver_")
                    else: target_task = task_map.get(old_task_display)

                    if target_task in tasks:
                        reward_vars.append(x[p, t_idx, target_task] * 2000000)

    # Scoring อื่นๆ
    for p in all_pharmacists:
        for t in range(15):
            for task in dispensing_tasks + ver_cpoe_tasks + ver_ps_tasks + ['Match_C', 'Match_C2']:
                match_var = model.NewBoolVar(f'p_{p}_{t}_{task}')
                model.AddImplication(match_var, x[p, t, task]); model.AddImplication(match_var, x[p, t+1, task])
                reward_vars.append(match_var * (500000 if task in dispensing_tasks else 150000))

    model.Maximize(sum(reward_vars))
    solver = cp_model.CpSolver()
    solver.parameters.num_workers = 8; solver.parameters.max_time_in_seconds = 60.0
    status = solver.Solve(model)

    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        schedule_data = []
        for p in all_pharmacists:
            row = {'ชื่อ/เวลา': p}
            for t in range(16):
                for task in tasks:
                    if solver.Value(x[p, t, task]) == 1:
                        if task == 'งานเฉพาะ': val = custom_dict_index.get((p, t), 'งานเฉพาะ')
                        elif task in ['นอกเวลา', 'ว่าง']: val = '-'
                        elif task == 'Match_C': val = 'Match + C'
                        elif task == 'Match_C2': val = 'Match + C2'
                        elif task == 'Ver_1': val = 'Ver 1 INC'
                        elif task == 'Ver_2': val = 'Ver 2/ปณ.'
                        elif task == 'Ver_3': val = 'Ver 3/A'
                        elif task.startswith('PS_'): val = 'Ver ' + task.replace('_', '')
                        elif task.startswith('จ่ายยา_'): val = task.replace('จ่ายยา_', 'จ่าย ')
                        else: val = task.replace('_', ' ')
                        row[time_slots[t]] = val
            schedule_data.append(row)
        df = pd.DataFrame(schedule_data)
        return df, "Success", ""
    return None, "Infeasible", "เงื่อนไขตึงเกินไปครับ"

def get_color_style(val):
    val_str = str(val); base = "text-align: center; " 
    if '/' in val_str and '-' in val_str and val_str[0].isdigit(): return base + 'background-color: #FFF2CC; font-weight: bold;' 
    if 'จ่าย ' in val_str: return base + 'background-color: #D5E8D4;' 
    if val_str == 'Matching': return base + 'background-color: #DAE8FC;' 
    if 'Match' in val_str: return base + 'background-color: #DAE8FC; color: red; font-weight: bold;' 
    if 'Ver PS' in val_str: return base + 'background-color: #E1D5E7;' 
    if 'Ver' in val_str: return base + 'background-color: #FFE6CC;' 
    if val_str == 'พัก': return base + 'background-color: #F8CECC;' 
    return base + 'background-color: #E6E6E6;' 

def build_html_table(df, selected_date, DAY_OF_WEEK):
    thai_date = get_thai_date(selected_date)
    def get_cell_style(val):
        val_str = str(val); bg, color, weight = "#E6E6E6", "black", "normal"
        if '/' in val_str and '-' in val_str and val_str and val_str[0].isdigit(): bg, weight = "#FFF2CC", "bold"
        elif 'จ่าย ' in val_str: bg = "#D5E8D4"
        elif val_str == 'Matching': bg = "#DAE8FC"
        elif 'Match' in val_str: bg, color, weight = "#DAE8FC", "red", "bold"
        elif 'Ver PS' in val_str: bg = "#E1D5E7"
        elif 'Ver' in val_str: bg = "#FFE6CC"
        elif val_str == 'พัก': bg = "#F8CECC"
        return f"background-color: {bg}; color: {color}; font-weight: {weight}; border: 1px solid black; padding: 4px; text-align: center; font-size: 17px; height: 50px;"

    cols = df.columns.tolist(); num = len(cols)
    html = f"<div id='capture-area' style='background-color: white; padding: 20px; display: inline-block; font-family: Sarabun;'><table style='border-collapse: collapse; width: 100%;'><tr><td colspan='{num}' style='text-align: center; font-size: 28px; font-weight: bold;'>ตารางปฏิบัติงานเภสัชกร ห้องยาชั้น 1 อาคารสมเด็จพระเทพรัตน์</td></tr><tr><td colspan='{num}' style='text-align: center; font-size: 22px; font-weight: bold;'>ประจำ{thai_date}</td></tr><tr>"
    for i, col in enumerate(cols): html += f"<th style='border: 1px solid black; padding: 6px; font-size: 19px;'>{col}</th>"
    html += "</tr>"
    for _, row in df.iterrows():
        html += "<tr>"
        for i, col in enumerate(cols):
            style = get_cell_style(row[col])
            if i == 0 or _ == len(df)-1: style = style.replace("font-weight: normal", "font-weight: bold")
            html += f"<td style='{style}'>{row[col]}</td>"
        html += "</tr>"
    html += "</table></div>"
    return html

# ==========================================
# 🖥️ UI - Streamlit
# ==========================================
st.set_page_config(page_title="Pharmacy AI Schedule", layout="wide")
st.markdown("<style>th {text-align: center !important;} hr {margin: 0.5rem 0;}</style>", unsafe_allow_html=True)

st.title("💊 จัดตารางปฏิบัติงานเภสัชกร ด้วย AI")
st.subheader("🏥 ห้องยาชั้น 1 อาคารสมเด็จพระเทพรัตน์")
st.markdown(f"<p style='color:gray;'>version 120.1 | เชื่อมต่อ Database: {'✅ พร้อม' if SHEETS_AVAILABLE else '❌ ไม่พร้อม'}</p>", unsafe_allow_html=True)

ft_list = ['เต้น', 'แอน', 'แม็ค', 'โบ้ท', 'ไม้เอก', 'กิ๊ฟ', 'ฟอร์จูน', 'มิ้ลค์', 'ริน', 'อ๊อฟฟี่', 'ออย', 'บี', 'มายด์', 'ขิม', 'บีม', 'มิ้น', 'ใบเตย', 'จีน่า', 'ปอนด์']
dropdown = ["ไม่มี"] + ft_list

if "schedule_df" not in st.session_state: st.session_state.schedule_df = None
if "ref_df" not in st.session_state: st.session_state.ref_df = None

with st.sidebar:
    st.markdown("## ⚙️ ตั้งค่าตาราง")
    st.subheader("📅 เลือกวันที่ประจำตาราง")
    sel_date = st.date_input("date", datetime.now().date(), label_visibility="collapsed")
    date_str = sel_date.strftime("%Y-%m-%d")
    IS_MWF = sel_date.weekday() in [0, 2, 4]
    DAY_OF_WEEK = 'Wed_Fri' if sel_date.weekday() in [2, 4] else 'Normal'
    st.success(f"✔️ ปรับตาราง {DAY_OF_WEEK} อัตโนมัติ")
    
    # ปุ่มดึงข้อมูลเดิม
    if st.button("🔍 ดึงตารางเดิมจาก Database (ถ้ามี)"):
        data = load_from_db(date_str)
        if data is not None:
            st.session_state.ref_df = data.set_index('ชื่อ/เวลา')
            st.success("พบตารางเดิม! ระบบเข้าสู่โหมด Rescheduling")
        else: st.error("ไม่พบข้อมูลของวันนี้")

    st.divider()
    st.subheader("🏖️ ผู้ที่ลาในวันนี้")
    leaves = {}
    with st.expander("คลิกเพื่อระบุคนลา"):
        for i in range(5):
            st.markdown(f"**คนที่ {i+1}**")
            c1, c2 = st.columns(2)
            p = c1.selectbox(f"ชื่อ {i}", dropdown, label_visibility="collapsed")
            t = c2.selectbox(f"ประเภท {i}", ["ทั้งวัน", "เช้า", "บ่าย"], label_visibility="collapsed")
            if p != "ไม่มี": leaves[p] = t

    st.subheader("🧑‍⚕️ เภสัชกร Part-time")
    pts = []
    with st.expander("คลิกเพื่อระบุ PT"):
        for i in range(5):
            st.markdown(f"**PT คนที่ {i+1}**")
            name = st.text_input(f"ชื่อ {i}", key=f"pt_n_{i}", placeholder="ชื่อ")
            cc1, cc2, cc3 = st.columns(3)
            s = cc1.selectbox(f"เริ่ม {i}", VALID_TIMES, index=0)
            e = cc2.selectbox(f"จบ {i}", VALID_TIMES, index=16)
            b = cc3.checkbox(f"พัก {i}", value=True)
            if name: pts.append({'name': name, 'start': s, 'end': e, 'has_break': b})

    st.subheader("📋 ภารกิจพิเศษ")
    cust = {}
    with st.expander("คลิกเพื่อระบุงานด่วน"):
        for i in range(20):
            st.markdown(f"**งานที่ {i+1}**")
            p = st.selectbox(f"คน {i}", dropdown, label_visibility="collapsed")
            n = st.text_input(f"งาน {i}", placeholder="ระบุชื่องาน")
            c1, c2 = st.columns(2)
            s = c1.selectbox(f"เริ่ม {i}s", VALID_TIMES, index=0)
            e = c2.selectbox(f"จบ {i}e", VALID_TIMES, index=2)
            if p != "ไม่มี" and n: cust[(p, s, e)] = n

    st.subheader("📌 ล็อกงานหลัก")
    fix_main = {}
    with st.expander("คลิกเพื่อล็อก"):
        opts = ['จ่าย 4', 'จ่าย 5', 'จ่าย 6', 'จ่าย 7', 'จ่าย 8', 'จ่าย 9', 'จ่าย 10', 'จ่าย 11', 'Ver 1 INC', 'Ver 2/ปณ.', 'Ver 3/A', 'Match + C', 'Matching']
        maps = {'จ่าย 4':'จ่ายยา_4','จ่าย 5':'จ่ายยา_5','จ่าย 6':'จ่ายยา_6','จ่าย 7':'จ่ายยา_7','จ่าย 8':'จ่ายยา_8','จ่าย 9':'จ่ายยา_9','จ่าย 10':'จ่ายยา_10','จ่าย 11':'จ่ายยา_11','Ver 1 INC':'Ver_1','Ver 2/ปณ.':'Ver_2','Ver 3/A':'Ver_3','Match + C':'Match_C','Matching':'Matching'}
        for i in range(10):
            p = st.selectbox(f"ล็อก {i}", dropdown, label_visibility="collapsed")
            n = st.selectbox(f"งาน {i}", ["เลือก"] + opts, label_visibility="collapsed")
            c1, c2 = st.columns(2)
            s = c1.selectbox(f"s {i}", VALID_TIMES, index=0); e = c2.selectbox(f"e {i}", VALID_TIMES, index=2)
            if p != "ไม่มี" and n != "เลือก": fix_main[(p, s, e)] = maps[n]

    st.subheader("🤒 คนป่วย")
    sick = []
    with st.expander("ระบุคนป่วย"):
        for i in range(3):
            p = st.selectbox(f"ป่วย {i}", dropdown); 
            if p != "ไม่มี": sick.append(p)

if st.button("🚀 เริ่มจัดตาราง / ซ่อมตารางด้วย AI", type="primary", use_container_width=True):
    with st.spinner("AI กำลังประมวลผล..."):
        res, status, msg = generate_schedule(DAY_OF_WEEK, leaves, cust, pts, {}, fix_main, sick, IS_MWF, st.session_state.ref_df)
        if status == "Success":
            st.session_state.schedule_df = res
            st.success("จัดตารางสำเร็จ!")
        else: st.error(msg)

if st.session_state.schedule_df is not None:
    df = st.session_state.schedule_df
    st.dataframe(df.style.applymap(get_color_style, subset=df.columns[1:]), use_container_width=True)
    
    c1, c2 = st.columns(2)
    with c1:
        if st.button("💾 บันทึกตารางลง Database (Google Sheets)", use_container_width=True):
            if save_to_db(df, date_str): st.balloons(); st.success("บันทึกสำเร็จ!")
            else: st.error("บันทึกล้มเหลว (กรุณาเช็กสิทธิ์การเข้าถึง Sheet)")
    with c2:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Schedule', startrow=2)
            # (โค้ดจัดรูปแบบ Excel เดิม...)
        st.download_button("📥 ดาวน์โหลด Excel", buffer.getvalue(), f"Schedule_{date_str}.xlsx", use_container_width=True)

    html = build_html_table(df, sel_date, DAY_OF_WEEK)
    file_png = f"Schedule_{date_str}.png"
    components.html(f"<!DOCTYPE html><html><head><meta charset='utf-8'><link href='https://fonts.googleapis.com/css2?family=Sarabun&display=swap' rel='stylesheet'><script src='https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js'></script></head><body><button style='width:100%; height:40px;' onclick='takeShot()'>📸 บันทึกเป็นรูปภาพ (PNG)</button><div id='capture-area-wrapper' style='position:absolute; left:-9999px;'>{html}</div><script>function takeShot(){{html2canvas(document.getElementById('capture-area'),{{scale:2}}).then(canvas=>{{let l=document.createElement('a');l.download='{file_png}';l.href=canvas.toDataURL();l.click();}})}}</script></body></html>", height=60)
